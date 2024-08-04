import openpyxl as opxl
import os


def validar_i_j(i, j, n):

    try:
        i = int(i)
        j = int(j)
        n = int(n)
    except ValueError:
        return False
    
    if j > n:
        return False
    if j < 1:
        return False
    if i < 1:
        return False
    if i + j > n + 1:
        return False

    return i, j


def validar_n(n_visitas):

    try:
        n_visitas = int(n_visitas)
    except ValueError:
        return False
    
    if not 10 <= n_visitas <= 1000000:
        return False
    
    return n_visitas


def acumular_probabilidades(probabilidades):
    acumulados = [probabilidades[0]]
    for i in range(1, len(probabilidades)):
        acumulados.append(probabilidades[i] + acumulados[i-1])
    return acumulados


def clasificar_numero_aleatorio(rnd, clases, probabilidad_x_clase):
    vector_probabilidades_acumuladas = acumular_probabilidades(probabilidad_x_clase)
    for i, prob in enumerate(vector_probabilidades_acumuladas):
        if rnd < prob:
            return clases[i]
        

def determinar_vpn_y_probabilidades(proyecto, inversion):
    proyectos = {
        'Proyecto A': {
            500000: ([0, 0.5, 1, 1.5], [0.25, 0.25, 0.4, 0.1]),
            1000000: ([0, 0.8, 1.2, 1.5, 2], [0.05, 0.2, 0.35, 0.25, 0.15]),
            1500000: ([0, 1, 1.8, 2.5, 3], [0.25, 0.25, 0.2, 0.2, 0.1]),
            2000000: ([0, 1, 2, 2.5, 3], [0.05, 0.15, 0.2, 0.3, 0.3])
        },
        'Proyecto B': {
            500000: ([0, 0.5, 1, 1.5], [0.25, 0.25, 0.4, 0.1]),
            1000000: ([0, 0.5, 1, 1.5, 2], [0.2, 0.2, 0.2, 0.2, 0.2]),
            1500000: ([0, 1, 1.8, 2.5, 3], [0.2, 0.2, 0.2, 0.2, 0.2]),
            2000000: ([0, 1, 2, 2.5, 3], [0.2, 0.15, 0.2, 0.1, 0.35])
        },
        'Proyecto C': {
            500000: ([0.2, 0.8, 1.4, 2], [0.5, 0.2, 0.2, 0.1]),
            1000000: ([0.5, 1, 1.5, 2], [0.3, 0.3, 0.2, 0.2]),
            1500000: ([0.05, 1, 1.8, 2.5, 3], [0.05, 0.25, 0.25, 0.25, 0.2]),
            2000000: ([1, 2, 2.5, 3], [0.25, 0.25, 0.25, 0.25])
        }
    }

    if inversion == 0:
        return None
    else:
        return proyectos[proyecto][inversion]


def formatear_nombre(inversiones):
    a, b, c = [str(inversion/1000000) for inversion in inversiones]
    return "Comb. " + a + "M - " + b + "M - " + c + "M"


def get_table(tablas, filepath="Simulacion.xlsx", auto_open=True):
    
    vector_estado, ultima_fila, inversiones = tablas[0]

    # Creamos el handler del workbook y la hoja
    wb = opxl.Workbook()
    ws = wb.active
    ws.title = formatear_nombre(inversiones)

    # Agregar primera fila de Headers
    ws.append(
        ['Iteracion', 'Rnd_A', 'VPN_A', "Rnd_B", 'VPN_B', "Rnd_C", 'VPN_C', "VPN", 
        'VPN_Acumulado', 'VPN_Promedio']
    )


    # Por cada fila a mostrar, redondear rnd's, y colocar -- si es None
    for fila in vector_estado:
        if fila[2] is None:
            fila[2] = "--"
        if fila[4] is None:
            fila[4] = "--"
        if fila[6] is None:
            fila[6] = "--"
        fila[1] = round(fila[1],4) if fila[1] is not None else "--"
        fila[3] = round(fila[3],4) if fila[3] is not None else "--"
        fila[5] = round(fila[5], 4) if fila[5] is not None else "--"
        fila[-1] = round(fila[-1], 4)

        ws.append(fila)
    
    ws.append(
        ['', '', '', "", '', "", '', "", '', '']
    )

    # Redondear la ultima iteracion y appendearla al sheet
    if ultima_fila[2] is None:
        ultima_fila[2] = "--"
    if ultima_fila[4] is None:
        ultima_fila[4] = "--"
    if ultima_fila[6] is None:
        ultima_fila[6] = "--"
    ultima_fila[1] = round(ultima_fila[1], 4) if ultima_fila[1] is not None else "--"
    ultima_fila[3] = round(ultima_fila[3], 4) if ultima_fila[3] is not None else "--"
    ultima_fila[5] = round(ultima_fila[5], 4) if ultima_fila[5] is not None else "--"
    ultima_fila[-1] = round(ultima_fila[-1], 4)

    ws.append(ultima_fila)


    # Por cada columna de la tabla
    for letra in "ABCDEFGHIJ":
        # Ajustar ancho de columnas
        if ws.column_dimensions[letra].has_style:
            ws.column_dimensions[letra] = None
        ws.column_dimensions[letra].width = 20
        
        for cell in ws[f"{letra}:{letra}"]:
            # Alinear celdas al centro
            cell.alignment = opxl.styles.Alignment(horizontal="center")

            # borders
            cell.border = opxl.styles.Border(
                        left=opxl.styles.Side(border_style="thin", color="000000"),
                        right=opxl.styles.Side(border_style="thin", color="000000"),
                        top=opxl.styles.Side(border_style="thin", color="000000"),
                        bottom=opxl.styles.Side(border_style="thin", color="000000")
                        )


        # Style the first row (Headers)
        ws[f"{letra}1"].font = opxl.styles.Font(bold=True)
        ws[f"{letra}1"].fill = opxl.styles.PatternFill(patternType="solid", fgColor="C4C2C1")
        
    for i in range(1, 15):

        vector_estado, ultima_fila, inversiones = tablas[i]
    
        # create new sheet and append the headers row
        new_sheet = wb.create_sheet(formatear_nombre(inversiones))
        
        # Agregar primera fila de Headers
        new_sheet.append(
            ['Iteracion', 'Rnd_A', 'VPN_A', "Rnd_B", 'VPN_B', "Rnd_C", 'VPN_C', "VPN", 
            'VPN_Acumulado', 'VPN_Promedio']
        )


        # Por cada fila a mostrar, redondear rnd's, y colocar -- si es None
        for fila in vector_estado:
            if fila[2] is None:
                fila[2] = "--"
            if fila[4] is None:
                fila[4] = "--"
            if fila[6] is None:
                fila[6] = "--"
            fila[1] = round(fila[1],4) if fila[1] is not None else "--"
            fila[3] = round(fila[3],4) if fila[3] is not None else "--"
            fila[5] = round(fila[5], 4) if fila[5] is not None else "--"
            fila[-1] = round(fila[-1], 4)

            new_sheet.append(fila)
        
        new_sheet.append(
            ['', '', '', "", '', "", '', "", '', '']
        )

        # Redondear la ultima iteracion y appendearla al sheet
        if ultima_fila[2] is None:
            ultima_fila[2] = "--"
        if ultima_fila[4] is None:
            ultima_fila[4] = "--"
        if ultima_fila[6] is None:
            ultima_fila[6] = "--"
        ultima_fila[1] = round(ultima_fila[1], 4) if ultima_fila[1] is not None else "--"
        ultima_fila[3] = round(ultima_fila[3], 4) if ultima_fila[3] is not None else "--"
        ultima_fila[5] = round(ultima_fila[5], 4) if ultima_fila[5] is not None else "--"
        ultima_fila[-1] = round(ultima_fila[-1], 4)

        new_sheet.append(ultima_fila)


        # Por cada columna de la tabla
        for letra in "ABCDEFGHIJ":
            # Ajustar ancho de columnas
            if new_sheet.column_dimensions[letra].has_style:
                new_sheet.column_dimensions[letra] = None
            new_sheet.column_dimensions[letra].width = 20
            
            for cell in new_sheet[f"{letra}:{letra}"]:
                # Alinear celdas al centro
                cell.alignment = opxl.styles.Alignment(horizontal="center")

                # borders
                cell.border = opxl.styles.Border(
                            left=opxl.styles.Side(border_style="thin", color="000000"),
                            right=opxl.styles.Side(border_style="thin", color="000000"),
                            top=opxl.styles.Side(border_style="thin", color="000000"),
                            bottom=opxl.styles.Side(border_style="thin", color="000000")
                            )


            # Style the first row (Headers)
            new_sheet[f"{letra}1"].font = opxl.styles.Font(bold=True)
            new_sheet[f"{letra}1"].fill = opxl.styles.PatternFill(patternType="solid", fgColor="C4C2C1")
        

    # Auto abrir el excel 
    if auto_open:
        os.startfile(filepath)
    

    wb.save(filepath)

if __name__ == "__main__":
    a = determinar_vpn_y_probabilidades("Proyecto A", 500000)
    b = determinar_vpn_y_probabilidades("Proyecto B", 500000)
    c = determinar_vpn_y_probabilidades("Proyecto C", 0)

    for proyecto in [a, b, c]:
        print(proyecto)
        print("\n")